import json

from openpyxl import load_workbook

from fairifier.services.fairds_excel_export import try_export_fairds_metadata_excel


def test_excel_export_prefers_isa_values_matrix_and_preserves_linkage_columns(tmp_path):
    metadata = {
        "isa_structure": {
            "observationunit": {
                "columns": ["observation unit name"],
                "rows": [{"observation unit name": "collapsed"}],
            }
        }
    }
    isa_values = {
        "observationunit": {
            "columns": [
                "observation unit identifier",
                "observation unit name",
                "study identifier",
            ],
            "rows": [
                {
                    "observation unit identifier": "ZYMO_LOG_mocktest_cwl",
                    "observation unit name": "ZYMO_LOG_mocktest_cwl",
                    "study identifier": "ZYMO_PRJEB29504",
                }
            ],
        }
    }
    (tmp_path / "metadata.json").write_text(json.dumps(metadata), encoding="utf-8")
    (tmp_path / "isa_values_json.json").write_text(json.dumps(isa_values), encoding="utf-8")

    out = try_export_fairds_metadata_excel(tmp_path, fair_ds_api_url="")

    assert out is not None
    wb = load_workbook(out, data_only=True)
    ws = wb["Observationunit"]
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    assert "observation unit identifier" in headers
    assert "study identifier" in headers
    row = {
        headers[col - 1]: ws.cell(2, col).value
        for col in range(1, ws.max_column + 1)
    }
    assert row["observation unit identifier"] == "ZYMO_LOG_mocktest_cwl"
    assert row["study identifier"] == "ZYMO_PRJEB29504"
