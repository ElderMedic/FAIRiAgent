"""Standalone FAIR-DS metadata Excel generator with multi-row support.

Generates ``metadata_fairds.xlsx`` from the ``isa_structure`` in
``metadata.json``.

Features
--------
* **Local generation** — no API dependency; can produce a fully formatted
  workbook from the ISA structure alone.
* **Multi-row support** — each ISA sheet gets one Excel row per entity
  (the ``columns`` + ``rows`` format).
* **Entity splitting** — semicolon-separated values, "Experiment N" patterns,
  and other heuristics expand single merged rows into per-entity rows.
* **Backward compatibility** — the flat ``fields`` list is still accepted.
* **FAIR-DS API integration** — when an API URL is provided and reachable,
  ``POST /api/isa`` supplies the column headers (and any data rows the
  server writes); the local filling logic adds the remaining rows.
* **Graceful degradation** — if ``openpyxl`` is not installed, the function
  falls back to the API path; if neither works it returns ``None``.
"""

from __future__ import annotations

import io
import json
import logging
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from fairifier.utils.entity_splitter import split_entities_in_isa_structure
from fairifier.utils.isa_order import ISA_LEVEL_ORDER

logger = logging.getLogger(__name__)

_XML_ILLEGAL_RE = re.compile(r"[\x00-\x08\x0B-\x0C\x0E-\x1F\uFFFE\uFFFF]")


def _excel_safe_value(value: Any) -> str:
    return _XML_ILLEGAL_RE.sub("", str(value))

# ── Row resolution ──────────────────────────────────────────────


def _resolve_rows(level_data: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Resolve data rows from an ISA sheet's *level_data*.

    Prefers ``rows`` (``columns`` + ``rows`` multi-row format); falls back
    to the flat ``fields`` list (backward-compatible single-row).

    Returns
    -------
    List[Dict[str, Any]]
        Each dict maps a **lowercase field name** to its string value.
        Empty list if no data is present.
    """
    rows: List[Dict[str, Any]] = level_data.get("rows", [])
    if rows:
        return rows

    fields: List[Dict[str, Any]] = level_data.get("fields", [])
    if fields:
        row: Dict[str, Any] = {}
        for f in fields:
            name = (f.get("field_name") or "").strip().lower()
            val = f.get("value")
            if name and val is not None:
                row[name] = str(val)
        if row:
            return [row]

    return []


# ── Header-column map builder ────────────────────────────────────


def _build_header_column_map(ws: "openpyxl.worksheet.worksheet.Worksheet") -> Dict[str, int]:
    """Build ``{lowercase_header: column_1_indexed}`` from Excel row 1."""
    header_col: Dict[str, int] = {}
    for col_idx in range(1, ws.max_column + 1):
        h = ws.cell(row=1, column=col_idx).value
        if h:
            header_col[str(h).strip().lower()] = col_idx
    return header_col


# ── Local Excel generation (no API needed) ───────────────────────


def _build_column_requirement_map(
    isa_structure: Dict[str, Any],
) -> Dict[str, Dict[str, str]]:
    """Return {sheet_name: {column_name: requirement}} from the ISA structure.

    Sources (in order): ``_field_definitions``, ``fields`` blocks, and
    ``column_metadata`` embedded in each sheet.

    requirement is one of ``MANDATORY``, ``RECOMMENDED``, or ``OPTIONAL``.
    """
    by_sheet: Dict[str, Dict[str, str]] = {}

    # 1. _field_definitions (from JSONGenerator)
    field_defs = isa_structure.get("_field_definitions", []) or []
    for fd in field_defs:
        if not isinstance(fd, dict):
            continue
        name = str(fd.get("field_name", "")).strip().lower()
        sheet = str(fd.get("isa_sheet", "")).strip().lower()
        req = str(fd.get("requirement", "")).strip().upper()
        if not name or not sheet or req not in ("MANDATORY", "RECOMMENDED", "OPTIONAL"):
            continue
        by_sheet.setdefault(sheet, {})[name] = req

    # 2. isa_structure.<sheet>.fields blocks
    for sheet_name, sheet_data in isa_structure.items():
        if not isinstance(sheet_data, dict):
            continue
        fields = sheet_data.get("fields", []) or []
        for f in fields:
            if not isinstance(f, dict):
                continue
            name = str(f.get("field_name", "")).strip().lower()
            req = str(f.get("requirement", "")).strip().upper()
            if not name or req not in ("MANDATORY", "RECOMMENDED", "OPTIONAL"):
                continue
            by_sheet.setdefault(sheet_name.lower(), {})[name] = req

    # 3. isa_structure.<sheet>.column_metadata (explicit annotation)
    for sheet_name, sheet_data in isa_structure.items():
        if not isinstance(sheet_data, dict):
            continue
        col_meta = sheet_data.get("column_metadata", {}) or {}
        if not isinstance(col_meta, dict):
            continue
        for col_name, req in col_meta.items():
            req = str(req).strip().upper()
            if req in ("MANDATORY", "RECOMMENDED", "OPTIONAL"):
                by_sheet.setdefault(sheet_name.lower(), {})[col_name.strip().lower()] = req

    return by_sheet


# ── requirement-label styling ───────────────────────────────────────
_REQ_FILLS: Dict[str, PatternFill] = {}
_REQ_FONTS: Dict[str, Font] = {}


def _requirement_label(requirement: str) -> str:
    return {"MANDATORY": "M", "RECOMMENDED": "R", "OPTIONAL": "O"}.get(requirement, "")


def _requirement_fill(requirement: str):
    if not _REQ_FILLS:
        from openpyxl.styles import PatternFill

        _REQ_FILLS.update({
            "MANDATORY": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
            "RECOMMENDED": PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid"),
            "OPTIONAL": PatternFill(start_color="F4F4F4", end_color="F4F4F4", fill_type="solid"),
        })
    return _REQ_FILLS.get(requirement)


def _requirement_font(requirement: str):
    if not _REQ_FONTS:
        from openpyxl.styles import Font

        _REQ_FONTS.update({
            "MANDATORY": Font(bold=True, color="BF8F00", size=9),
            "RECOMMENDED": Font(bold=False, color="38761D", size=9),
            "OPTIONAL": Font(bold=False, color="999999", size=9),
        })
    return _REQ_FONTS.get(requirement)


def _generate_xlsx_local(
    isa_structure: Dict[str, Any],
) -> Optional[bytes]:
    """Generate a complete ``metadata_fairds.xlsx`` from *isa_structure*.

    Uses **openpyxl** only.  Returns raw ``.xlsx`` bytes or ``None`` if
    the library is unavailable or an error occurs.
    """
    try:
        import openpyxl
        from openpyxl.styles import (
            Alignment,
            Border,
            Font,
            PatternFill,
            Side,
        )
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.warning("openpyxl not installed; cannot generate Excel locally")
        return None

    # ── Style definitions ───────────────────────────────────────
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(vertical="top", wrap_text=False)
    thin_side = Side(style="thin")
    thin_border = Border(
        left=thin_side, right=thin_side, top=thin_side, bottom=thin_side
    )
    # ── end style definitions ───────────────────────────────────

    # ── Build requirement lookup ─────────────────────────────────
    column_req = _build_column_requirement_map(isa_structure)

    wb = openpyxl.Workbook()
    # Remove the auto-created empty sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    isa_level_order = list(ISA_LEVEL_ORDER)
    isa_descriptions = {
        "investigation": "Investigation-level metadata (project info)",
        "study": "Study-level metadata (experimental design)",
        "assay": "Assay-level metadata (measurement details)",
        "sample": "Sample-level metadata (biological material)",
        "observationunit": "ObservationUnit-level metadata (individual observations)",
        "help": "Help — workbook overview",
    }

    # ── ISA data sheets ─────────────────────────────────────────
    for level_name in isa_level_order:
        level_data = isa_structure.get(level_name, {})
        if not isinstance(level_data, dict):
            continue

        columns: List[str] = level_data.get("columns", [])
        if not columns:
            logger.debug("Skipping '%s' sheet: no columns defined", level_name)
            continue

        rows = _resolve_rows(level_data)
        ws = wb.create_sheet(title=level_name.capitalize())

        # Row 1: column headers (with requirement badge)
        req_map = column_req.get(level_name, {})
        for col_idx, col_name in enumerate(columns, start=1):
            key = col_name.strip().lower()
            req = req_map.get(key, "")
            label = _requirement_label(req)
            display = f"{col_name}  ({label})" if label else col_name
            cell = ws.cell(row=1, column=col_idx, value=display)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # Row 2+: data rows
        for row_idx, row_data in enumerate(rows, start=2):
            for col_idx, col_name in enumerate(columns, start=1):
                key = col_name.strip().lower()
                value = row_data.get(key)
                if value is not None:
                    cell = ws.cell(
                        row=row_idx, column=col_idx,
                        value=_excel_safe_value(value),
                    )
                    cell.alignment = cell_align
                    cell.border = thin_border

        # Column widths (auto-fit, clamped 10–60)
        for col_idx, col_name in enumerate(columns, start=1):
            col_letter = get_column_letter(col_idx)
            lengths = [len(str(col_name))]
            for row_data in rows:
                val_str = str(row_data.get(col_name.strip().lower(), ""))
                lengths.append(len(val_str))
            width = min(max(max(lengths) + 2, 10), 60)
            ws.column_dimensions[col_letter].width = width

        # Freeze header row
        ws.freeze_panes = "A2"

        logger.debug(
            "Generated '%s' sheet: %d columns x %d rows",
            level_name,
            len(columns),
            len(rows),
        )

    # ── Help sheet ──────────────────────────────────────────────
    ws_help = wb.create_sheet(title="Help", index=len(wb.sheetnames))
    help_lines = [
        ["FAIR-DS Metadata Excel Export"],
        [""],
        [
            "This workbook contains metadata organised by ISA-TAB levels: "
            "Investigation, Study, ObservationUnit, Sample, Assay."
        ],
        [""],
        ["Sheets:"],
        ["  Investigation"] + ["Project-level metadata (title, authors, contacts)"],
        ["  Study"] + ["Study-level metadata (experimental design, description)"],
        ["  ObservationUnit"] + ["Individual observation metadata"],
        ["  Sample"] + ["Sample-level metadata (biological material, environment)"],
        ["  Assay"] + ["Assay-level metadata (measurements, protocols, files)"],
        [""],
        [
            "Each sheet has one header row followed by one data row per entity. "
            "Fields present in the column header but absent in a particular row "
            "are considered 'not specified'."
        ],
        [""],
        ["Generated by FAIRiAgent"],
    ]
    for row_idx, row_data in enumerate(help_lines, start=1):
        for col_idx, value in enumerate(row_data, start=1):
            ws_help.cell(row=row_idx, column=col_idx, value=value)
    ws_help.column_dimensions["A"].width = 100

    try:
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
    except Exception as exc:
        logger.warning("Failed to save Excel workbook: %s", exc)
        return None


# ── Post-process: fill rows into API-generated workbook ─────────


def _fill_missing_data_rows(
    xlsx_bytes: bytes,
    isa_structure: Dict[str, Any],
) -> bytes:
    """Post-process Excel: fill data rows for ISA sheets the API left empty.

    The FAIR-DS ``POST /api/isa`` endpoint writes data rows only for
    ``investigation`` and ``study``.  This function fills the remaining
    sheets from the ``columns`` + ``rows`` (or legacy ``fields``) data in
    ``isa_structure`` — **one Excel row per entity**.

    Returns
    -------
    bytes
        The (possibly modified) ``.xlsx`` bytes.
    """
    try:
        import openpyxl
    except ImportError:
        return xlsx_bytes

    try:
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    except Exception:
        return xlsx_bytes

    isa_level_names = set(ISA_LEVEL_ORDER)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if sheet_name.lower() not in isa_level_names:
            continue
        # Only fill if the sheet has headers but no data rows yet.
        if ws.max_row > 1:
            continue

        level_data = isa_structure.get(sheet_name.lower())
        if not isinstance(level_data, dict):
            continue

        rows = _resolve_rows(level_data)
        if not rows:
            continue

        header_col = _build_header_column_map(ws)
        if not header_col:
            continue

        for row_data in rows:
            for key in row_data.keys():
                normalized = str(key).strip().lower()
                if not normalized or normalized in header_col:
                    continue
                col_idx = ws.max_column + 1
                ws.cell(row=1, column=col_idx, value=normalized)
                header_col[normalized] = col_idx

        total_filled = 0
        for row_idx, row_data in enumerate(rows, start=2):
            filled_in_row = 0
            for key, value in row_data.items():
                col = header_col.get(key.lower())
                if col is not None and value is not None:
                    ws.cell(row=row_idx, column=col, value=_excel_safe_value(value))
                    filled_in_row += 1
            total_filled += filled_in_row

        if total_filled:
            logger.debug(
                "Filled %d values across %d row(s) in '%s' sheet",
                total_filled,
                len(rows),
                sheet_name,
            )

    try:
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
    except Exception:
        return xlsx_bytes


# ── Public API ──────────────────────────────────────────────────


def try_export_fairds_metadata_excel(
    output_dir: Path,
    *,
    fair_ds_api_url: Optional[str] = None,
) -> Optional[Path]:
    """Write ``metadata_fairds.xlsx`` when ``metadata.json`` has fillable ISA data.

    Workflow
    --------
    1. Read ``metadata.json``, extract ``isa_structure``.
    2. Apply entity-splitting heuristics to expand single merged rows.
    3. If a FAIR-DS API URL is available **and** the server is reachable,
       call ``POST /api/isa`` to obtain a workbook with column headers.
       Then fill data rows from the (split) ISA structure.
    4. If the API is unavailable or not configured, generate the workbook
       entirely locally with ``openpyxl``.
    5. Return the path to the written ``.xlsx`` file, or ``None`` on
       failure / missing data.

    Parameters
    ----------
    output_dir : Path
        Directory containing ``metadata.json`` (and where the ``.xlsx``
        will be written).
    fair_ds_api_url : str, optional
        Override the configured FAIR-DS API URL.  ``None`` (default) uses
        ``config.fair_ds_api_url``.

    Returns
    -------
    Path or None
        Absolute path to the written ``metadata_fairds.xlsx`` file.
    """
    from ..config import config
    from ..output_paths import (
        FAIRDS_METADATA_EXCEL_FILENAME,
        resolve_metadata_output_read_path,
    )
    from .fair_data_station import FAIRDataStationClient

    url = config.fair_ds_api_url if fair_ds_api_url is None else fair_ds_api_url

    meta_path = resolve_metadata_output_read_path(Path(output_dir))
    if not meta_path:
        logger.debug(
            "FAIR-DS Excel export skipped: no metadata.json in %s",
            output_dir,
        )
        return None

    try:
        with open(meta_path, encoding="utf-8") as fh:
            data: Dict[str, Any] = json.load(fh)
    except (OSError, json.JSONDecodeError) as exc:
        logger.warning(
            "FAIR-DS Excel export skipped: cannot read %s: %s",
            meta_path,
            exc,
        )
        return None

    isa_structure = data.get("isa_structure")
    if not isinstance(isa_structure, dict) or not isa_structure:
        logger.debug(
            "FAIR-DS Excel export skipped: missing or empty isa_structure"
        )
        return None

    # Prefer dedicated matrix file; fall back to isa_structure.
    isa_values_path = Path(output_dir) / "isa_values_json.json"
    try:
        with open(isa_values_path, encoding="utf-8") as fh:
            isa_values = json.load(fh)
        if isinstance(isa_values, dict):
            fill_structure = split_entities_in_isa_structure(isa_values)
        else:
            fill_structure = split_entities_in_isa_structure(isa_structure)
    except (OSError, json.JSONDecodeError):
        fill_structure = split_entities_in_isa_structure(isa_structure)

    # Check for fillable content.
    has_content = False
    for _level, block in fill_structure.items():
        if isinstance(block, dict) and (
            block.get("fields") or block.get("rows")
        ):
            has_content = True
            break
    if not has_content:
        logger.debug(
            "FAIR-DS Excel export skipped: no fields/rows under isa_structure"
        )
        return None

    # ── Step 1: Entity splitting ───────────────────────────────
    # Defensive: even if the JSON generator already split entities,
    # re-apply to catch any remaining semicolons in merged rows.
    isa_structure = split_entities_in_isa_structure(isa_structure)

    # ── Step 2: Generate Excel ─────────────────────────────────
    xlsx_bytes: Optional[bytes] = None
    api_used = False

    if url and str(url).strip():
        base = str(url).strip().rstrip("/")
        client = FAIRDataStationClient(base)
        if client.is_available():
            try:
                xlsx_bytes = client.generate_excel_from_isa_structure(
                    fill_structure
                )
                api_used = True
            except Exception as exc:
                logger.warning(
                    "FAIR-DS API Excel generation failed: %s", exc
                )
        else:
            logger.debug(
                "FAIR-DS API not reachable at %s; falling back to local generation",
                base,
            )
    else:
        logger.debug(
            "FAIR-DS API URL not set; using local Excel generation"
        )

    if xlsx_bytes is not None and api_used:
        # The API returned headers (and perhaps a few data rows);
        # fill the remaining data locally.
        xlsx_bytes = _fill_missing_data_rows(xlsx_bytes, fill_structure)
    elif xlsx_bytes is None:
        # API was not used or failed — generate entirely locally.
        xlsx_bytes = _generate_xlsx_local(fill_structure)

    if xlsx_bytes is None:
        logger.warning(
            "FAIR-DS Excel export failed: no xlsx bytes generated"
        )
        return None

    out = Path(output_dir) / FAIRDS_METADATA_EXCEL_FILENAME
    try:
        out.write_bytes(xlsx_bytes)
    except OSError as exc:
        logger.warning(
            "FAIR-DS Excel export could not write %s: %s",
            out,
            exc,
        )
        return None

    logger.info(
        "FAIR-DS metadata Excel written: %s (%s bytes)",
        out.name,
        len(xlsx_bytes),
    )
    return out


__all__ = [
    "try_export_fairds_metadata_excel",
    "split_entities_in_isa_structure",
    "_generate_xlsx_local",
    "_fill_missing_data_rows",
    "_resolve_rows",
]
